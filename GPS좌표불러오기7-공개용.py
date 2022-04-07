import openpyxl  #pip install openpyxl  <-로 먼저 설치. 엑셀파일 XLSX을 불러오고 쓸 수 있는 것
import os #첨부파일의 유무 확인용. 
import sys 
import datetime

import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
#from math import ceil
from time import sleep #끝나고 바로 닫으면 메시지를 못읽으니 잠시 읽을시간을 주는 용도
import requests #pip install requests

#실행파일로 만들려면 pyinstaller --console --onefile GPS좌표불러오기7.py
# pip install pyinstaller

#http://api.vworld.kr/req/address?service=address&request=getAddress&version=2.0&crs=epsg:4326&point=127.06689704458516,37.37267617393529&format=json&type=both&zipcode=true&simple=false&key=A073D350-BF66-3A4E-8E9F-92FDFF7D65BC

def myf브이월드인증키불러오기():
    import dropbox #pip install dropbox, 드랍박스에서 브이월드 인증키를 받아옴. 인증키 갱신때문에 프로그램을 분기별로 만들기 귀찮아서 요렇게 받아옴. 추후 갱신이 필요하면 드롭박스 파일만 갈으면 됨.
    ACCESS_TOKEN = "드롭박스 엑세스 키를 넣어주세요"
    dbx = dropbox.Dropbox(ACCESS_TOKEN)
    with open("C:\\Users\\Public\\apiTokenKey.txt", "wb") as f:
        metadata, res = dbx.sharing_get_shared_link_file("드롭박스에서 브이월드인증키를 저장한 파일링크를 걸어주세요")
        f.write(res.content)
    try:
        with open("C:\\Users\\Public\\apiTokenKey.txt", "r", encoding='UTF-8') as f:
            myv주소변환인증키 = f.readline().rstrip("\n") #첫줄이 인증키
            ##print(myv주소변환인증키)
    except:
        myv주소변환인증키 = '이도 저도 모르겠다면 브이월드에서 발급받으신 인증키를 여기다 적어주세요'
    return myv주소변환인증키


def myf브이월드좌표로동리검색(위도,경도,myv주소변환인증키):
    try:    
        #아래 것으로 조회를 하고
        r = requests.get('http://api.vworld.kr/req/address?service=address&request=getAddress&version=2.0&crs=epsg:4326&point='+경도+','+위도+'&format=json&type=both&zipcode=true&simple=false&key='+myv주소변환인증키) 
        result = r.json() 

        print('---구조소로 일단 변환하고 새주소로 변환할게요--> '+ result['response']['result'][0]['text']+"\n")
        return result['response']['result'][0]['text']   #어짜피 저장하는거. 그냥 일반주소로 저장함
    except:
        return ''


def myf브이월드구주소로새주소및좌표검색(검색할지역,검색할번지,myv주소변환인증키): #data.go.kr 함수가 너무 구리고 인증키도 3개월마다 갱신하라고 해서 갈아탐
    print("\n---검색중입니다. ＿φ(￣ー￣ )   :  ", 검색할지역 )

    try:    
        #아래 것으로 조회를 하고
        r = requests.get('http://api.vworld.kr/req/search?service=search&request=search&version=2.0&type=address&category=parcel&format=json&key='+myv주소변환인증키+'&query='+검색할지역+" "+검색할번지) 
        result = r.json() 

        #print(result['response']['result']['items'][0]) #중간에 [0] 이 부분은 검색결과가 2개인 경우도 있기 때문, 첫번째가 가장 정확한 검색값이니 0으로 해줘야함 
        새주소 =   result['response']['result']['items'][0]['address']['road'] 
        구주소 =   result['response']['result']['items'][0]['address']['parcel'] 
        우편번호 = result['response']['result']['items'][0]['address']['zipcode']
        위도 = str(result['response']['result']['items'][0]['point']['y']) 
        경도 = str(result['response']['result']['items'][0]['point']['x'])
        print("------구주소로 완료\n\n")
    

    except:  #아파트 신축같이 기존 주소가 없는 곳의 경우 수동으로 입력해 줘야함
        구주소 = ""
        새주소 = ""
        우편번호 = ""
        위도 = ""
        경도 = ""
        '''
        구주소 = 검색할지역+" "+검색할번지; 새주소 = "새주소없는곳";  우편번호 = "우편번호없는곳"; 위도 = ""; 경도 = ""
        import webbrowser
        import time
        #url = "https://www.google.co.kr/maps/@37.4562325,126.8956891,10z?hl=ko"
        url = "https://www.google.co.kr/maps/search/"+검색할지역+" "+검색할번지
        webbrowser.open(url)
        temp = input("새주소 및 GPS좌표 검색이 안되네요. \n\n열린 구글지도 현장위치에서 마우스 오른쪽 버턴 -> \n가장 위 좌표를 눌러 복사 후(좌표가 안보이시면 한번 더 눌러주세요), \n여기 옆에서 마우스오른쪽버턴을 눌러, 좌표를 붙여 넣어주세요. \n좌표위치 : ")
        
        
        try:
            #위도 = str(round(float(temp[0:int(temp.find(","))].strip()),6))
            #경도 = str(round(float(temp[int(temp.find(","))+1:].strip()),6))
            위도 = str(temp[0:int(temp.find(","))].strip())
            경도 = str(temp[int(temp.find(","))+1:].strip())
        except:
            print("\n..\n....\n......\n.........\n위도 경도가 뭔가 이상하네요... 다시 확인해주세요\n")
        '''

    #os.remove("apiTokenKey.txt")
    #print(구주소, 새주소, 우편번호, 위도, 경도)
    return 구주소, 새주소, 우편번호, 위도, 경도  



def myf브이월드새주소로구주소및좌표검색(검색내용, 구색맞추기용,myv주소변환인증키): #data.go.kr 함수가 너무 구리고 인증키도 3개월마다 갱신하라고 해서 갈아탐
    print("---새주소인가요? 새주소로 검색 중입니다. ＿φ(￣ー￣ )   :  ", 검색내용 )
    try:    
        #아래 것으로 조회를 하고
        r = requests.get('http://api.vworld.kr/req/search?service=search&request=search&version=2.0&type=address&category=road&format=json&key='+myv주소변환인증키+'&query='+검색내용) 
        result = r.json() 

        #print(result['response']['result']['items'][0]) #중간에 [0] 이 부분은 검색결과가 2개인 경우도 있기 때문, 첫번째가 가장 정확한 검색값이니 0으로 해줘야함 
        새주소 =   result['response']['result']['items'][0]['address']['road'] 
        구주소 =   result['response']['result']['items'][0]['address']['parcel'] 
        우편번호 = result['response']['result']['items'][0]['address']['zipcode']
        위도 = str(result['response']['result']['items'][0]['point']['y']) 
        경도 = str(result['response']['result']['items'][0]['point']['x'])
        print("------새주소로 완료\n\n")
    

    except:  #아파트 신축같이 기존 주소가 없는 곳의 경우 수동으로 입력해 줘야함
        구주소 = ""
        새주소 = ""
        우편번호 = ""
        위도 = ""
        경도 = ""
        print("------오류가 있네요: "+검색내용)

    return 구주소, 새주소, 우편번호, 위도, 경도  


def myf주소로검색후자료넣기(주소):
    myv구새우위경 = myf브이월드구주소로새주소및좌표검색(주소,"",myv주소변환인증키)
    if myv구새우위경[0] == "":
        myv구새우위경 = myf브이월드새주소로구주소및좌표검색(주소,"",myv주소변환인증키)

    sheet.cell(i,2).value = (myv구새우위경[3]) #위도
    sheet.cell(i,3).value = (myv구새우위경[4]) #경도
    sheet.cell(i,4).value = (myv구새우위경[0]) #구주소
    sheet.cell(i,5).value = (myv구새우위경[1]) #새주소
    sheet.cell(i,6).value = (myv구새우위경[2]) #우편번호
    #sheet.cell(i,7).value = "검색안됨" if (myv구새우위경[0]) == "" else ""
    book.save(myPath+'/Out.xlsx')


# 엑셀파일 열기
try:
    book = openpyxl.load_workbook(sys.argv[1]) #마우스로 엑셀파일을 실행파일위에 끌어다 놓았을때
    myv엑셀파일 = sys.argv[1]
    myPath = os.path.dirname(sys.argv[1])   #엑셀파일이 있는 위치
except:
    
    print("신청서 엑셀파일을 선택해주세요(*˘◡˘*)")
    root = tk.Tk()
    root.withdraw()
    temp = filedialog.askopenfilename(title = "신청서 엑셀파일을 선택해주세요(*˘◡˘*)") #엑셀파일을 수동으로 찾음
    book = openpyxl.load_workbook(temp)
    myv엑셀파일 = temp
    myPath = os.path.dirname(temp)   #엑셀파일이 있는 위치




# 맨 앞의 시트 추출하기
sheet = book.worksheets[0]
myv주소변환인증키 = myf브이월드인증키불러오기()

for i in range(2,30000):   #엑셀에서 3만건만 조회함. vworld에서 지원하는 하루 요청 건수가 하루 3만건.                    
    if sheet.cell(i,2).value != None and sheet.cell(i,3).value != None:
        print(str(i)+"번째 열을 시작합니다.φ(．．;)")
        임시주소 = myf브이월드좌표로동리검색(sheet.cell(i,2).value,sheet.cell(i,3).value,myv주소변환인증키)
        myf주소로검색후자료넣기(임시주소)
    elif sheet.cell(i,1).value == None:
        pass
    else:
        print(str(i)+"번째 열을 시작합니다.φ(．．;)")
        myf주소로검색후자료넣기(sheet.cell(i,1).value)

now = datetime.datetime.now()
#book.save('./Out-'+now.strftime('%H-%M-%S')+'.xlsx')
print("\n\n완료했습니다.  Out.xlsx 파일을 열어보세요(*˘◡˘*)")

os.startfile(myPath)


sleep(10)


